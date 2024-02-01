import os
from openpyxl import load_workbook


def test_xlsx(create_zip):
    xlsx_path = os.path.join(create_zip, 'file_example_XLSX_50.xlsx')
    work_book = load_workbook(xlsx_path)
    sheet = work_book.active

    assert sheet is not None

    expected_rows = 51
    expected_columns = 8
    max_row = sheet.max_row
    max_column = sheet.max_column

    assert max_row == expected_rows
    assert max_column == expected_columns

    title = sheet.cell(row=1, column=5).value
    assert title == "Country"

    work_book.close()
